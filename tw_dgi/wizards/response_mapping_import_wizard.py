# -*- coding: utf-8 -*-
from odoo import models, fields, api
from odoo.exceptions import UserError
import csv
import base64
import io


class TWResponseMappingImportWizard(models.TransientModel):
    _name = "tw.response.mapping.import.wizard"
    _description = "Import Response Mapping from Excel/CSV"
    
    endpoint_id = fields.Many2one(
        "tw.endpoint.configuration",
        string="Endpoint",
        required=True,
        readonly=True
    )
    
    file = fields.Binary(
        string="Excel File",
        help="Upload Excel (.xlsx) file with response mappings"
    )
    
    filename = fields.Char(string="Filename")
    
    def action_download_template(self):
        """Download Excel template"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise UserError("openpyxl library not installed. Please install it first.")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Response Mapping Template"
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Define headers (reordered: field_type first)
        headers = [
            "sequence", "json_path", "field_type", "target_model", "target_field",
            "is_required", "default_value", "notes",
            "relation_model", "relation_lookup_field", "relation_search_domain", "validation"
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add sample data (reordered, with validation example)
        sample_data = [
            [5, "data.0.statusCode", "validation", "", "", "TRUE", "", "Validation-only: check status without saving (target_model & target_field optional)", "", "", "", ""],
            [10, "status", "direct", "your.model", "api_status", "TRUE", "", "Direct mapping: save value to field", "", "", "", ""],
            [20, "data.0.id", "direct", "your.model", "external_id", "TRUE", "", "Direct mapping: external ID", "", "", "", ""],
            [30, "data.0.category_id", "relation", "your.model", "category_id", "FALSE", "", "Relational: lookup existing record", "res.partner.category", "external_id", "[('active','=',True)]", ""],
        ]
        
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Encode to base64
        file_content = base64.b64encode(output.read())
        
        # Create attachment
        attachment = self.env["ir.attachment"].create({
            "name": "response_mapping_template.xlsx",
            "datas": file_content,
            "type": "binary",
        })
        
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "new",
        }
    
    def action_import(self):
        """Import mappings from uploaded Excel file"""
        self.ensure_one()
        
        if not self.file:
            raise UserError("Please upload a file first.")
        
        if not self.filename or not self.filename.endswith(".xlsx"):
            raise UserError("Please upload an Excel file (.xlsx)")
        
        try:
            file_data = base64.b64decode(self.file)
            mappings, parse_errors = self._parse_xlsx(file_data)
            
            # Track results
            created_count = 0
            updated_count = 0
            skipped_count = 0
            errors = parse_errors.copy()  # Start with parse errors
            
            for idx, mapping_data in enumerate(mappings, 1):
                if not mapping_data:
                    skipped_count += 1
                    errors.append(f"Row {idx}: Skipped (empty or invalid)")
                    continue
                
                try:
                    mapping_data["endpoint_id"] = self.endpoint_id.id
                    
                    # Check if mapping with same json_path already exists
                    existing = self.env["tw.mapping.response"].search([
                        ("endpoint_id", "=", self.endpoint_id.id),
                        ("json_path", "=", mapping_data["json_path"])
                    ], limit=1)
                    
                    if existing:
                        # Update existing mapping
                        existing.write(mapping_data)
                        updated_count += 1
                    else:
                        # Create new mapping
                        self.env["tw.mapping.response"].create(mapping_data)
                        created_count += 1
                        
                except Exception as e:
                    skipped_count += 1
                    errors.append(f"Row {idx} ({mapping_data.get('json_path', 'Unknown')}): {str(e)}")
                    raise UserError(f"Error importing row {idx}: {str(e)}")
            
            # Build detailed message
            total = created_count + updated_count
            message = f"Import completed: {total} mappings processed"
            
            if created_count > 0:
                message += f"\n✓ {created_count} created"
            if updated_count > 0:
                message += f"\n✓ {updated_count} updated"
            if skipped_count > 0:
                message += f"\n✗ {skipped_count} skipped/failed"
            
            if errors:
                message += "\n\nErrors:\n" + "\n".join(errors[:10])  # Show first 10 errors
                if len(errors) > 10:
                    message += f"\n... and {len(errors) - 10} more errors"
            
            notification_type = "success" if total > 0 else "warning"
            
            return {
                "type": "ir.actions.client",
                "tag": "display_notification",
                "params": {
                    "title": "Import Results",
                    "message": message,
                    "type": notification_type,
                    "sticky": True,  # Keep notification visible
                }
            }
            
        except Exception as e:
            raise UserError(f"Error importing response mappings: {str(e)}")
            raise UserError(f"Error importing file: {str(e)}")
    
    def _parse_xlsx(self, file_data):
        """Parse Excel file and return (mappings, errors)"""
        try:
            import openpyxl
        except ImportError:
            raise UserError("openpyxl library not installed. Please use CSV format.")
        
        workbook = openpyxl.load_workbook(io.BytesIO(file_data))
        sheet = workbook.active
        
        headers = [cell.value for cell in sheet[1]]
        
        mappings = []
        errors = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = dict(zip(headers, row))
            mapping, error = self._row_to_mapping(row_dict, row_idx)
            
            if mapping:
                mappings.append(mapping)
            elif error:
                errors.append(error)
        
        return mappings, errors
    
    def _row_to_mapping(self, row, row_number):
        """Convert row dict to mapping data
        Returns: (mapping_data, error_message)
        """
        # Skip completely empty rows
        if not any(row.values()):
            return None, None
            
        if not row.get("json_path"):
            return None, f"Row {row_number}: Missing 'json_path' column"
        
        def to_bool(val):
            if isinstance(val, bool):
                return val
            if isinstance(val, str):
                return val.strip().upper() in ("TRUE", "1", "YES", "Y")
            return False
        
        def safe_strip(val):
            """Safely strip string, handle None"""
            if val is None:
                return ""
            return str(val).strip()
        
        json_path = safe_strip(row.get("json_path"))
        field_type = safe_strip(row.get("field_type")) or "direct"
        
        # target_model required for direct/relation, optional for validation
        target_model_name = safe_strip(row.get("target_model"))
        target_model = None
        
        if field_type in ("direct", "relation"):
            if not target_model_name:
                return None, f"Row {row_number} ({json_path}): Missing 'target_model' for field_type '{field_type}'"
                
            target_model = self.env["ir.model"].search([
                ("model", "=", target_model_name)
            ], limit=1)
            
            if not target_model:
                return None, f"Row {row_number} ({json_path}): Model '{target_model_name}' not found in system. Please install the module or check model name."
        
        elif field_type == "validation" and target_model_name:
            # Optional for validation, but validate if provided
            target_model = self.env["ir.model"].search([
                ("model", "=", target_model_name)
            ], limit=1)
        
        if not json_path:
            return None, f"Row {row_number}: Empty 'json_path' after processing"
        
        # target_field required for direct/relation, optional for validation
        target_field = safe_strip(row.get("target_field"))
        
        if field_type in ("direct", "relation") and not target_field:
            return None, f"Row {row_number} ({json_path}): Missing 'target_field' for field_type '{field_type}'"

        mapping_data = {
            "sequence": int(row.get("sequence", 10)),
            "json_path": json_path,
            "field_type": field_type,
            "is_required": to_bool(row.get("is_required", False)),
        }
        
        # target_model optional for validation type
        if target_model:
            mapping_data["target_model_id"] = target_model.id
        
        # target_field optional for validation type
        if target_field:
            mapping_data["target_field"] = target_field
        
        # Optional fields - only add if not empty
        default_value = safe_strip(row.get("default_value"))
        if default_value:
            mapping_data["default_value"] = default_value
        
        notes = safe_strip(row.get("notes"))
        if notes:
            mapping_data["notes"] = notes
        
        # Lookup relation model if provided
        relation_model_name = safe_strip(row.get("relation_model"))
        if relation_model_name:
            relation_model = self.env["ir.model"].search([
                ("model", "=", relation_model_name)
            ], limit=1)
            
            if relation_model:
                mapping_data["relation_model_id"] = relation_model.id
            else:
                raise UserError(f"Row {row_number} ({json_path}): Relation model '{relation_model_name}' not found")
        
        relation_lookup = safe_strip(row.get("relation_lookup_field"))
        if relation_lookup:
            mapping_data["relation_lookup_field"] = relation_lookup
        
        relation_domain = safe_strip(row.get("relation_search_domain"))
        if relation_domain:
            mapping_data["relation_search_domain"] = relation_domain
        
        validation = safe_strip(row.get("validation"))
        if validation:
            mapping_data["validation"] = validation
        
        return mapping_data, None
