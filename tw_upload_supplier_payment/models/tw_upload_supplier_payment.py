import base64
import openpyxl
import xlrd
import logging

from datetime import date
from io import BytesIO
from itertools import groupby
from operator import itemgetter

from odoo import api, fields, models, _
from odoo.exceptions import ValidationError as Warning

_logger = logging.getLogger(__name__)

# * Columns mapping
COLS = {
    'payment_method': 0,
    'no_invoice': 1,
    'amount': 2,
    'division': 3,
    'entry_name': 4,
    'inter_branch_code': 5,
    'partner_code': 6,
    'source_document': 7,
    'amount_unreconciled': 8,
    'writeoff_account': 9,
    'writeoff_amount': 10,
}
MANDATORY_COLS = {
    'inter_branch_code': 'Bayar Untuk',
    'partner_code': 'Supplier',
    'payment_method': 'Payment Method',
    'amount': 'Total',
    'amount_unreconciled': 'Total',
    'entry_name': 'Journal Item',
}

class TwUploadSupplierPayment(models.TransientModel):
    _name = "tw.upload.supplier.payment"
    _description = "Upload Supplier Payment"

    file = fields.Binary('File to Upload')
    file_name = fields.Char()
    message = fields.Text()

    def _read_cell_value(self, value, force_zero_padding=False, use_fraction=False):
        if value:
            if use_fraction:
                value = str(value).replace('.0','').replace(',0','')
                return float(value)
            # ? Fix for phone numbers that supposedly has leading zero but miswritten so it detected as float
            if isinstance(value, float) and force_zero_padding:
                value = '0'+str(int(value))
            else:
                try:
                    value = str(value) if not isinstance(value, str) else value
                except UnicodeEncodeError as err:
                    _logger.error(err.args)
                    value = str(value.encode('utf-8'))
                    
                while (value[0] == ' ' or value[len(value)-1] == ' '):
                    if value[0] == ' ':
                        value = value[1:]
                    if value[len(value)-1] == ' ':
                        value = value[:len(value)-1]
        return value
    
    def _get_excel_column(self,n):
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string

    def action_upload(self):
        if not self.file:
            raise Warning("File Excel harus terisi!")
        data = base64.decodebytes(self.file)
        
        # do checking for xls or xlsx file type
        ext = self.file_name.split('.')[-1]
        if ext == 'xls':
            excel = xlrd.open_workbook(file_contents=data)
            sheet = excel.sheet_by_index(0)
        
        elif ext == 'xlsx':
            excel = openpyxl.load_workbook(BytesIO(data), data_only=True)
            sheet = excel.active
            
        else:
            raise Warning('TIDAK dapat upload file dengan ekstensi %s' % ext)

        err = ''
        fail = 0
        success_ids = []
        data = []

        iteration = range(1, sheet.nrows) if ext == 'xls' else range(1, sheet.max_row)
        x = 1 if ext == 'xlsx' else 0
        for i in iteration:
            # * Collect all columns data
            row = {
                'payment_method': self._read_cell_value(sheet.cell(i + x, COLS['payment_method'] + x).value),
                'no_invoice': self._read_cell_value(sheet.cell(i + x, COLS['no_invoice'] + x).value),
                'amount': self._read_cell_value(sheet.cell(i + x, COLS['amount'] + x).value, use_fraction=True),
                'division': self._read_cell_value(sheet.cell(i + x, COLS['division'] + x).value),
                'entry_name': self._read_cell_value(sheet.cell(i + x, COLS['entry_name'] + x).value),
                'inter_branch_code': self._read_cell_value(sheet.cell(i + x, COLS['inter_branch_code'] + x).value),
                'partner_code': self._read_cell_value(sheet.cell(i + x, COLS['partner_code'] + x).value),
                'source_document': self._read_cell_value(sheet.cell(i + x, COLS['source_document'] + x).value),
                'amount_unreconciled': self._read_cell_value(sheet.cell(i + x, COLS['amount_unreconciled'] + x).value, use_fraction=True),
                'writeoff_account': self._read_cell_value(sheet.cell(i + x, COLS['writeoff_account'] + x).value),
                'writeoff_amount': self._read_cell_value(sheet.cell(i + x, COLS['writeoff_amount'] + x).value, use_fraction=True),
            }
            # If all row is None skip the iteration
            if not all([False for v in row.values() if v is None]):
                continue

            # * Check for missing / invalid mandatory values
            for item in MANDATORY_COLS:
                if not row[item] or row[item] == '0':
                    col = self._get_excel_column(COLS[item]+1)
                    cell = col+str(i+1)
                    err += "%s di cell %s TIDAK boleh kosong!" % (MANDATORY_COLS[item], cell)
                    fail += 1
                    _logger.error(err)
            
            inter_branch = self.env['res.company'].search([('code', '=', row['inter_branch_code'])], limit=1)
            if not inter_branch:
                raise Warning("Dealer dengan Code '%s' TIDAK ditemukan!" % row['inter_branch_code'])
            
            partner = self.env['res.partner'].search([('code', '=', row['partner_code'])], limit=1)
            if not partner:
                raise Warning("Supplier/Customer dengan code '%s' TIDAK ditemukan!" % row['partner_code'])
            
            journal = self.env['account.journal'].search([('name', '=', row['payment_method'])], limit=1)
            if not journal:
                raise Warning("Payment Method '%s' TIDAK ditemukan!" % row['payment_method'])
            
            row['beneficiary_company_id'] = inter_branch.id
            row['partner_id'] = partner.id
            row['partner_name'] = partner.name
            row['journal_id'] = journal.id
            data.append(row)

        # * Check HO in Res Company
        ho_id = self.env['res.company'].search([('code', '=', 'HHO')], limit=1)
        if not ho_id:
            raise Warning("No HO (Head Office) found in your Company data")
        
        grouper = itemgetter('beneficiary_company_id', 'partner_id', 'journal_id', 'division')
        sorted_data = sorted(data, key=grouper)
        for key, group in groupby(sorted_data, grouper):
            vals = dict(zip(['beneficiary_company_id', 'partner_id', 'journal_id', 'division'], key))
            journal_id = self.env['account.journal'].browse(vals.get('journal_id'))
            amount = 0
            wo_ids = []
            vals_line_dr_ids = []
            vals_line_wo_ids = []
            for g in group:
                move_line_id = self.env['account.move.line'].search([('move_id.name', '=', g['entry_name']),
                                                                     ('name', '=', g['source_document']),
                                                                     ('partner_id', '=', g['partner_id']),
                                                                     ('credit', '>', 0)])
                if not move_line_id:
                    err += "Move Line dengan Entry Name '%s' dan Partner '%s' TIDAK ditemukan!" % (g['entry_name'], g['partner_name'])
                    fail += 1
                    continue

                reconciled = move_line_id._check_reconciled()
                if reconciled:
                    err += reconciled
                    fail += 1
                    continue

                remaining_amount = g['amount_unreconciled']
                currency = self.env.user.company_id.currency_id or journal_id.company_id.currency_id
                
                if move_line_id.currency_id and currency == move_line_id.currency_id:
                    amount_original = abs(move_line_id.amount_currency)
                    amount_unreconciled = abs(move_line_id.amount_residual_currency)
                else:
                    #always use the amount booked in the company currency as the basis of the conversion into the voucher currency
                    amount_original = currency.round(move_line_id.credit or move_line_id.debit or 0.0)
                    amount_unreconciled = currency.round(abs(move_line_id.amount_residual))
                    
                vals_line_dr_ids.append([0, 0, {
                    'name':move_line_id.move_id.name,
                    'move_line_id':move_line_id.id,
                    'amount_original': amount_original,
                    'amount': move_line_id and min(abs(remaining_amount), amount_unreconciled) or 0.0,
                    'date_original':move_line_id.date,
                    'date_due':move_line_id.date_maturity,
                    'amount_unreconciled': amount_unreconciled,
                    'account_id': move_line_id.account_id.id,
                    'type': 'cr' if move_line_id.debit else 'dr',
                    'currency_id': move_line_id.currency_id and move_line_id.currency_id.id or move_line_id.company_id.currency_id.id,
                    'partner_id': vals.get('partner_id'),
                }]) 

                if g.get('writeoff_account') and g.get('writeoff_amount'):
                    writeoff_account = self.env['account.account'].search([('code', '=', g['writeoff_account'])])
                    if writeoff_account:
                        wo_ids.append({
                            'name': writeoff_account.name,
                            'account_id': writeoff_account.id,
                            'type': 'wo',
                            'amount': g['writeoff_amount'],
                            'partner_id': vals.get('partner_id'),
                        })

                amount += g.get('amount_unreconciled', 0) + g.get('writeoff_amount', 0)

            vals['company_id'] = ho_id.id
            vals['amount'] = amount
            vals['date'] = date.today()
            vals['partner_type'] = 'supplier'
            vals['type'] = 'supplier_payment'
            vals['payment_type'] = 'outbound'
            vals['currency_id'] = journal_id.company_id.currency_id.id or self.env.user.company_id.currency_id.id
            vals['account_id'] = journal_id.default_credit_account_id.id or journal_id.default_debit_account_id.id
            vals['is_round'] = True
            vals['line_dr_ids'] = vals_line_dr_ids
            
            # grouping and summarize wo by its account
            if wo_ids:
                wo_grouper = itemgetter('name', 'account_id')
                vals_wo = sorted(wo_ids, key=wo_grouper)
                for k, wo in groupby(vals_wo, wo_grouper):
                    wo_val = dict(zip(['name', 'account_id'], k))
                    wo_val.update({
                        'type': 'wo',
                        'amount': sum([w.get('amount') for w in wo]),
                        'partner_id': vals.get('partner_id'),
                    })
                    vals_line_wo_ids.append([0, 0, wo_val])

                vals['line_wo_ids'] = vals_line_wo_ids

            try:
                pv = self.env['tw.account.payment'].create(vals)
                success_ids += pv.ids
            except Exception as msg:
                err += '%s\n' % str(msg)
                fail += 1

        if fail or err:
            raise Warning(err)
        
        tree_id = self.env.ref('tw_payment.tw_supplier_payment_list_view').id
        form_id = self.env.ref('tw_payment.tw_account_payment_form_view').id
        action_id = self.env.ref('tw_payment.tw_supplier_payment_action').id
        return {
            'type': 'ir.actions.act_window',
            'name': 'Supplier Payments',
            'res_model': 'tw.account.payment',
            'view_mode': 'list,form',
            'domain': [('type', '=', 'supplier_payment'), ('id', 'in', success_ids),('payment_type','=','outbound'),('partner_type','=','supplier')],
            'context': {
                'default_payment_type': 'outbound',
                'default_type': 'supplier_payment',
                'default_partner_type': 'supplier',
                'search_default_outbound_filter': 1,
                'default_move_journal_types': ('bank', 'cash'),
                'display_account_trust': True,
                'action_id': action_id,
            },
            'views': [(tree_id, 'list'), (form_id, 'form')],
            'search_view_id': self.env.ref('tw_payment.tw_account_payment_search_view').id,
            'target': 'current'
        }
    
    def action_download_format_file(self):
        format = self.env['tw.format.upload'].suspend_security().search([('name', '=', 'upload supplier payment'), ('active', '=', True)], limit=1)
        if format:
            return {
                'type': 'ir.actions.act_url',
                'name': 'contract',
                'url': f'/web/content/tw.format.upload/{format.id}/file_format_show/{format.filename_upload_format}'
            }
        else:
            raise Warning(_("Sorry, the format for upload supplier payment is not available yet. Please contact the Helpdesk."))
