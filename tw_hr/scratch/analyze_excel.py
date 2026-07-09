import openpyxl
import json

file_path = '/Users/liong/Documents/Code/18_teds/tw_hr/migration/xlsx_import/migration_employee.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

header = [cell.value for cell in sheet[1]]

found_row = None
for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    row_dict = dict(zip(header, row))
    if any(str(v) == 'DMX2412003' for v in row):
        found_row = i
        break

if found_row:
    print(f"Found NIP 'DMX2412003' at Row {found_row}")
    data_dict = dict(zip(header, [cell.value for cell in sheet[found_row]]))
    print(json.dumps(data_dict, indent=2, default=str))
else:
    print("Could not find 'DDI2006001' in the first sheet.")
    # Check other sheets?
    if len(wb.sheetnames) > 1:
        print(f"Sheets: {wb.sheetnames}")
