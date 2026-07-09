# 1: imports of python lib
from datetime import timedelta, datetime
import base64
from io import BytesIO
import openpyxl

from odoo import models, fields, api
from odoo.exceptions import UserError as Warning



class UploadMasterMargin(models.TransientModel):
    _name = "tw.upload.master.margin"
    _description = "Upload Master Margin"

    def _get_default_date(self):
        return self.env['res.company'].get_default_date()

    file = fields.Binary('File')
    date = fields.Date('Tanggal',readonly=True,default=_get_default_date)
    state_x = fields.Selection([('choose','choose'),('get','get')],default=lambda self:self._context.get('default_state_x','choose'))
    
    def action_download_format_file(self):
        # Create Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Master Target Margin"
        
        # Headers matching user requested format
        headers = ["Branch", "Job", "Series", "Manufacture Year", "Cash", "Credit"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        
        # Data: Active Series for both 'sales' and 'sco'
        series_ids = self.env['product.series'].sudo().search([('active', '=', True)])
        current_year = datetime.now().year
        company_code = self.env.company.code or ''
        
        row = 2
        for job_code in ['sales', 'sco']:
            for idx, series in enumerate(series_ids):
                # Pre-fill data
                if idx == 0:
                    ws.cell(row=row, column=1, value=company_code)
                    ws.cell(row=row, column=2, value=job_code)
                
                ws.cell(row=row, column=3, value=series.name)
                ws.cell(row=row, column=4, value=current_year)
                ws.cell(row=row, column=5, value=20000)
                ws.cell(row=row, column=6, value=250000)
                row += 1
            
        # Save to buffer
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        
        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': 'format_master_target_margin.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(data),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }

    
    def _create_record(self, vals, code, job):
        company_id = self.env['res.company'].sudo().search([('code', '=', code)],limit=1)
        vals.update({'company_id': company_id.id}) 
        self.env['tw.master.target.margin'].suspend_security().create(vals)
        vals = {}
    
    def action_import(self):
        if not self.file:
            raise Warning("Silahkan input file terlebih dahulu.")
        
        data = base64.b64decode(self.file)
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        sh = wb.active
        
        warning_note = ''
        rows = list(sh.iter_rows(min_row=2, values_only=True))
        
        # loop untuk error checking
        for idx, values in enumerate(rows, start=2):
            if not values or not any(v is not None for v in values):
                continue
            
            if values[0]:
                company_id = self.env['res.company'].sudo().search([('code', '=', str(values[0]))], limit=1)
                if not company_id: 
                    warning_note += 'Baris ke %s kode dealer %s tidak ditemukan\n' % (idx, values[0])
            
            job = values[1]
            if job and job not in ('sales', 'sco', 'sc'):
                warning_note += 'Baris ke %s kategori job %s tidak ditemukan\n' % (idx, values[1])
            
            if values[2]:
                series = self.env['product.series'].sudo().search([('name', '=', str(values[2]))], limit=1)
                if not series:
                    warning_note += 'Baris ke %s series %s tidak ditemukan\n' % (idx, values[2])
        
            if values[3]:
                tahun = values[3]
                if not isinstance(tahun, int) or len(str(tahun)) != 4:
                    warning_note += 'Baris ke %s tahun %s tidak valid\n' % (idx, values[3])
        
        if warning_note:
            raise Warning(warning_note)

        vals = {}
        line_ids = []
        company_id = ''
        job = ''
        
        for idx, values in enumerate(rows, start=2):
            if not values or not any(v is not None for v in values):
                continue
                
            series = self.env['product.series'].sudo().search([('name', '=', str(values[2]))], limit=1)
            
            # Logic grouping by job/company
            if values[1] and job:
                # Save previous group
                vals.update({
                    'job': job,
                    'target_margin_line_ids': line_ids
                })
                self._create_record(vals, company_id, job)
                
                # Reset for next group
                job = values[1]
                line_ids = []
            
            if values[0]:
                company_id = str(values[0])
            
            if not job and values[1]:
                job = values[1]

            line_ids.append([0, 0, {
                'series_id': series.id if series else False,
                'year': str(values[3]) if values[3] is not None else False,
                'cash': values[4] or 0,
                'credit': values[5] or 0,
            }])

        # Final group
        if job and line_ids:
            vals.update({
                'job': job,
                'target_margin_line_ids': line_ids
            })
            self._create_record(vals, company_id, job)
                    
        return {
            'type': 'ir.actions.act_window',
            'name': 'Master Margin',
            'res_model': 'tw.master.target.margin',
            'view_type': 'list',
            'view_mode': 'list,form',
            'target': 'current',
        }
