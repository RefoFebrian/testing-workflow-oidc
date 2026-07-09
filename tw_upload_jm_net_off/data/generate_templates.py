"""
One-time script to generate XLSX example templates.
Run from the data/ directory: python generate_templates.py
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Shared helpers ────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
KEY_FILL     = PatternFill("solid", fgColor="D6E4F0")   # light blue
KEY_FONT     = Font(name="Calibri", bold=True, size=11)
DATA_FONT    = Font(name="Calibri", size=11)
OPT_FILL     = PatternFill("solid", fgColor="FFF2CC")   # light yellow (optional cols)
THIN         = Side(style="thin", color="BFBFBF")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _set(ws, row, col, value, font=None, fill=None, align=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:           cell.font          = font
    if fill:           cell.fill          = fill
    if align:          cell.alignment     = align
    if number_format:  cell.number_format = number_format
    cell.border = BORDER
    return cell


def _col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── 1. Journal Memorial template ──────────────────────────────────────────────

def make_jm_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Upload JM"

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    # ── Column headers (row 1) — 12 columns A-L ───────────────────────────────
    JOURNAL_ITEM_FILL = PatternFill("solid", fgColor="E2EFDA")  # light green
    col_headers = [
        "Branch",                                    # A — header-level
        "Periode",                                   # B — header-level
        "description",                               # C — header-level
        "Auto Reverse ?",                            # D — header-level
        "Division",                                  # E — header-level
        "Partner Code",                              # F — header-level (optional)
        "Journal memorial line/Branch",              # G — per-line
        "Journal memorial line/Account",             # H — per-line
        "journal_memorial_line/type",                # I — per-line
        "Journal memorial line/Amount",              # J — per-line
        "Journal memorial line/Reconcile",          # K — per-line: existing move line name (optional)
        "Journal memorial line/Asset Code",          # L — per-line (optional)
    ]
    opt_cols        = {12}   # L=Asset Code is optional (1-based); K=Reconcile gets its own fill
    journal_item_col = 11        # K column (1-based)
    for c, h in enumerate(col_headers, start=1):
        if c == journal_item_col:
            fill = JOURNAL_ITEM_FILL
            font = Font(name="Calibri", bold=True, color="000000", size=10)
        elif c in opt_cols:
            fill = OPT_FILL
            font = Font(name="Calibri", bold=True, color="000000", size=10)
        else:
            fill = HEADER_FILL
            font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        _set(ws, 1, c, h, font=font, fill=fill, align=center)

    # ── Example data rows ─────────────────────────────────────────────────────
    # Tuple: (Branch, Periode, Desc, AutoRev, Division, Partner, LineBranch, Account, Type, Amount, Asset, JournalItem)
    rows = [
        # Group 1: lines reference existing journal items for Net Off (col K)
        ("HHO", "2026/002", "reclass Accrue Blind Bonus MD ke rtl Lampung", "FALSE", "Unit", "MML", "MML", "21629907", "Dr", 61418960.25, "JM/HO/24/08/00006", ""),
        ("",    "",         "",                                              "",      "",     "",    "DLH", "21629907", "Cr",  4920570,    "IC/23/05/00066",    ""),
        ("",    "",         "",                                              "",      "",     "",    "DLK", "21629907", "Cr",  6691290,    "",                  ""),
        ("",    "",         "",                                              "",      "",     "",    "DLX", "21629907", "Cr",  2637187,    "",                  ""),
        # Group 2: no journal item references — JM stays in draft
        ("HBD", "2026/002", "Penyesuaian Sparepart Feb 2026",               "TRUE",  "Sparepart", "", "HBD", "11001001", "Dr", 10000000, "",                  ""),
        ("",    "",         "",                                              "",      "",     "",    "HBD", "21001001", "Cr", 10000000,    "",                  ""),
    ]
    for r_off, row in enumerate(rows):
        r = 2 + r_off
        for c, val in enumerate(row, start=1):
            fmt = '#,##0.00' if c == 10 else None
            fill = JOURNAL_ITEM_FILL if (c == 11 and val) else None
            _set(ws, r, c, val, font=DATA_FONT, fill=fill,
                 align=right if c == 10 else left, number_format=fmt)

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [8, 10, 42, 13, 12, 14, 14, 22, 8, 18, 22, 22]
    for c, w in enumerate(widths, start=1):
        _col_width(ws, c, w)

    ws.freeze_panes = "A2"

    # ── Note row ──────────────────────────────────────────────────────────────
    note_row = len(rows) + 3
    ws.cell(row=note_row, column=1,
            value="* Kolom F (Partner Code) bersifat opsional. "
                  "Kolom A-F hanya diisi pada baris pertama setiap grup JM. "
                  "Type: Dr = Debit, Cr = Credit. "
                  "Kolom K (Reconcile) diisi dengan nama Journal Item yang sudah ada untuk direkonsiliasi. "
                  "Kolom L (Asset Code) bersifat opsional. "
                  "Aktifkan 'Auto Net Off' pada wizard jika kolom K diisi."
                  ).font = Font(italic=True, color="7F7F7F", size=10)

    path = os.path.join(OUTPUT_DIR, "template_upload_journal_memorial.xlsx")
    wb.save(path)
    print(f"Saved: {path}")


# ── 2. Net Off template ───────────────────────────────────────────────────────

def make_net_off_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Upload Net Off"

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    # ── Column headers (row 1) ────────────────────────────────────────────────
    col_headers = [
        "Branch Code", "Partner Code", "Description",
        "Account Code", "Account Move Line", "Debit", "Credit",
    ]
    opt_cols = {2}   # Partner Code is optional
    for c, h in enumerate(col_headers, start=1):
        fill = OPT_FILL if c in opt_cols else HEADER_FILL
        font = Font(name="Calibri", bold=True, color="000000" if c in opt_cols else "FFFFFF", size=11)
        _set(ws, 1, c, h, font=font, fill=fill, align=center)

    # ── Example data rows ─────────────────────────────────────────────────────
    # Group 1: two lines, branch/partner/description only on first row
    rows = [
        ("AY", "DAV0003002", "Net off JM/HO/24/08/00006", "212102", "JM/HO/24/08/00006", 37_425_595.83, 0),
        ("",   "",           "",                           "212102", "IC/23/05/00066",    0,             37_425_595.83),
        # Group 2 example
        ("AY", "DAV0004001", "Net off IC/25/01/00012",    "212102", "IC/25/01/00012",    12_000_000,    0),
        ("",   "",           "",                           "212102", "JM/HO/25/01/00003", 0,            12_000_000),
    ]
    for r_off, row in enumerate(rows):
        r = 2 + r_off
        for c, val in enumerate(row, start=1):
            fmt = '#,##0.00' if c in (6, 7) else None
            _set(ws, r, c, val, font=DATA_FONT, align=right if c in (6, 7) else left, number_format=fmt)

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [14, 14, 32, 16, 24, 18, 18]
    for c, w in enumerate(widths, start=1):
        _col_width(ws, c, w)

    ws.freeze_panes = "A2"

    # ── Note row ──────────────────────────────────────────────────────────────
    note_row = 7
    ws.cell(row=note_row, column=1,
            value="* Branch Code, Partner Code, dan Description hanya diisi pada baris pertama setiap grup. "
                  "Baris berikutnya dalam grup yang sama biarkan kosong.").font = Font(italic=True, color="7F7F7F", size=10)

    path = os.path.join(OUTPUT_DIR, "template_upload_net_off.xlsx")
    wb.save(path)
    print(f"Saved: {path}")


if __name__ == "__main__":
    make_jm_template()
    make_net_off_template()
    print("Done.")
