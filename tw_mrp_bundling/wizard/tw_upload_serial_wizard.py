# -*- coding: utf-8 -*-

import base64
import io
import openpyxl
import xlsxwriter
from odoo import api, fields, models, _
from odoo.exceptions import UserError

class TwUploadSerialWizard(models.TransientModel):
    _name = 'tw.upload.serial.wizard'
    _description = 'Upload Serial Number Wizard'

    production_id = fields.Many2one('mrp.production', string='Production Order', required=True)
    file_data = fields.Binary(string='File', attachment=False)
    file_name = fields.Char(string='File Name')

    def action_download_template(self):
        self.ensure_one()
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Serial Numbers')
        
        # Header formatting
        header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3'})
        worksheet.write(0, 0, 'Product Name', header_format)
        worksheet.write(0, 1, 'Serial Number', header_format)
        
        # Adjust column widths
        worksheet.set_column(0, 0, 40)
        worksheet.set_column(1, 1, 30)

        # Pre-populate product name if it's a bundling order
        if self.production_id.order_type == 'bundling':
            move_raw_with_lot_ids = self.production_id.move_raw_ids.filtered(lambda ln: ln.product_id.tracking == 'serial')
            if move_raw_with_lot_ids:
                product_name = move_raw_with_lot_ids[0].product_id.display_name
                qty = int(self.production_id.product_qty)
                for i in range(qty):
                    worksheet.write(i + 1, 0, product_name)

        workbook.close()
        
        file_data = base64.b64encode(output.getvalue())
        self.write({
            'file_data': file_data,
            'file_name': f'Template_Serial_{self.production_id.name or "New"}.xlsx'
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/?model=tw.upload.serial.wizard&id={self.id}&field=file_data&download=true&filename={self.file_name}',
            'target': 'self',
        }

    def action_upload(self):
        self.ensure_one()
        if not self.file_data:
            raise UserError(_("Please upload an Excel file."))
            
        try:
            file_content = base64.b64decode(self.file_data)
            wb = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True)
            ws = wb.active
        except Exception as e:
            raise UserError(_("Invalid file format. Please upload a valid Excel (.xlsx) file. Error: %s") % str(e))
            
        if self.production_id.order_type != 'bundling':
            raise UserError(_("This feature is currently only supported for bundling orders."))
            
        move_raw = self.production_id.move_raw_ids.filtered(lambda ln: ln.product_id.tracking == 'serial')
        if not move_raw:
            raise UserError(_("No serial-tracked component found in this production order."))
        if len(move_raw) > 1:
            raise UserError(_("Only one raw material with lot/serial number is allowed for bundling production."))
            
        move_raw = move_raw[0]
        expected_product = move_raw.product_id
        
        # Parse serials
        serials_to_assign = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Stop if the row is completely empty
            if not any(row):
                continue
                
            product_name = str(row[0] or '').strip()
            serial_number = str(row[1] or '').strip()
            
            if not product_name and not serial_number:
                continue
                
            if not serial_number:
                raise UserError(_("Row %s: Serial Number is missing.") % i)
                
            if product_name and product_name != expected_product.display_name:
                raise UserError(_("Row %(row)s: Product Name '%(input)s' does not match the expected component '%(expected)s'.") % {
                    'row': i,
                    'input': product_name,
                    'expected': expected_product.display_name
                })
                
            # Check availability
            lot = self.env['stock.lot'].search([
                ('name', '=', serial_number),
                ('product_id', '=', expected_product.id),
                ('company_id', '=', self.production_id.company_id.id)
            ], limit=1)
            
            if not lot:
                raise UserError(_("Row %s: Serial Number '%s' does not exist in the system for product '%s'.") % (i, serial_number, expected_product.display_name))
                
            # Use standard get_stock_available for the lot if possible, but actually we need to check if the specific lot is in the location.
            # Easiest way is to check the quants directly for that lot and location.
            quant = self.env['stock.quant'].search([
                ('lot_id', '=', lot.id),
                ('location_id', '=', move_raw.location_id.id),
                ('quantity', '>', 0)
            ],limit=1)
            if quant.reserved_quantity > 0:
                reserved_by = quant.reservation_ids[0].picking_id.name if (quant.reservation_ids and quant.reservation_ids[0].picking_id) else 'Other'
                raise UserError(_("Row %s: Serial Number '%s' is reserved for '%s'.") % (i, serial_number, reserved_by))
                
            available_qty = quant.quantity - quant.reserved_quantity
            if available_qty <= 0:
                raise UserError(_("Row %s: Serial Number '%s' is not available in location '%s'.") % (i, serial_number, move_raw.location_id.display_name))
                
            serials_to_assign.append(lot)
            
        if not serials_to_assign:
            raise UserError(_("No valid serial numbers found in the uploaded file."))
            
        # Optional: Warn if count mismatch, but we will just replace whatever is there.
        # Clear existing move lines for this move
        move_raw.move_line_ids.unlink()
        
        # Assign the new serial numbers
        move_line_vals = []
        for lot in serials_to_assign:
            move_line_vals.append((0, 0, {
                'product_id': expected_product.id,
                'lot_id': lot.id,
                'quantity': 1,
                'product_uom_id': expected_product.uom_id.id,
                'location_id': move_raw.location_id.id,
                'location_dest_id': move_raw.location_dest_id.id,
                'company_id': move_raw.company_id.id,
                'picking_id': move_raw.picking_id.id if move_raw.picking_id else False,
            }))
            
        move_raw.write({'move_line_ids': move_line_vals})
        move_raw.picked = True
        
        # Return an action to close the wizard and refresh the view
        return {'type': 'ir.actions.act_window_close'}
