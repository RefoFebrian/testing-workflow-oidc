import openpyxl
import base64
from io import BytesIO

from datetime import date, datetime, timedelta
from dateutil import parser, relativedelta

from odoo import models, fields, api, _, Command
from odoo.exceptions import UserError as Warning
from odoo.tools import SQL

class UploadProfitBeforeTax(models.TransientModel):
    _name = "tw.upload.profit.before.tax"
    _description = "Upload Profit Before Tax"

    def _get_default_date(self):
        return self.env['res.company'].get_default_date()

    def _get_default_datetime(self): 
        return datetime.now() + timedelta(hours=7)

    file = fields.Binary('File')
    date = fields.Date('Tanggal', readonly=True, default=_get_default_date)
    state_x = fields.Selection([('choose','choose'), ('get','get')], default=lambda self:self._context.get('default_state_x','choose'))
    messages = fields.Text("message", readonly=True)
    
    def action_download_format_file(self):
        format = self.env['tw.format.upload'].suspend_security().search([('name','=', 'upload pbt'),('active','=',True)],limit=1)
        if format:
            return {
                'type': 'ir.actions.act_url',
                'name': 'contract',
                'url': '/web/content/tw.format.upload/%s/file_format_show/%s?download=true' % (format.id, format.filename_upload_format)
            }
        else:
            raise Warning('Maaf, format belum tersedia. Silahkan Hubungi Helpdesk.')
    
    def _create_record(self, vals, code, job):
        company_id = self.env['res.company'].search([('code', '=', code)], limit=1)
        vals.update({'company_id': company_id.id})
        self.env['tw.profit.before.tax'].suspend_security().create(vals)
        vals = {}

    def isinstance_check(self, value):
        if isinstance(value, str) and value == '':
            return True
        else:
            return False
    
    def action_import(self):
        if not self.file:
            raise Warning('Silakan input file terlebih dahulu.')

        # deleted soon, just for testing
        list_rows = []
        
        data = base64.b64decode(self.file)
        excel = openpyxl.load_workbook(BytesIO(data), data_only=True)
        sheet = excel.worksheets[0]
        
        warning_header = ''
        
        # error checking untuk header
        branch_code_value = sheet.cell(row=1, column=3).value
        branch_obj = self.env['res.company'].search([('code', '=', branch_code_value)],limit=1)
        if not branch_obj:
            warning_header += 'Baris ke %s Kode Cabang %s tidak ada\n' % (1, branch_code_value)
        
        period_value = sheet.cell(row=2, column=3).value
        if self.isinstance_check(period_value):
            warning_header += 'Baris ke %s data PERIODE tidak ada\n' % (2)

        opex_value = sheet.cell(row=3, column=3).value
        if self.isinstance_check(opex_value):
            warning_header += 'Baris ke %s data OPEX AVG tidak ada\n' % (3)

        total_unit_value = sheet.cell(row=6, column=3).value
        if self.isinstance_check(total_unit_value):
            warning_header += 'Baris ke %s data TOTAL UNIT tidak ada\n' % (5)
        
        total_cash_value = sheet.cell(row=7, column=3).value
        if self.isinstance_check(total_cash_value):
            warning_header += 'Baris ke %s data TOTAL CASH tidak ada\n' % (6)
        
        total_credit_value = sheet.cell(row=8, column=3).value
        if self.isinstance_check(total_credit_value):
            warning_header += 'Baris ke %s data TOTAL CREDIT tidak ada\n' % (6)
        
        total_net_margin_value = sheet.cell(row=9, column=3).value
        if self.isinstance_check(total_net_margin_value):
            warning_header += 'Baris ke %s data TOTAL SISA MARGIN tidak ada\n' % (6)
        
        total_refund = sheet.cell(row=10, column=3).value
        if self.isinstance_check(total_refund):
            warning_header += 'Baris ke %s data TOTAL REFUND tidak ada\n' % (7)
        
        pbt_propose_value = sheet.cell(row=11, column=3).value
        if self.isinstance_check(pbt_propose_value):
            warning_header += 'Baris ke %s data PBT PROPOSE tidak ada\n' % (7)

        if warning_header:
            raise Warning('Tidak dapat diproses, karena ada kesalahan.\nDetail:\n{}'.format(warning_header))
        
        dict_data = {'header': {}}
        
        company_id = branch_obj.id
        dict_data['header'].update({'kode_cabang': company_id})

        period_str = str(period_value) if period_value else ""
        new_period_value = [item.strip() for item in period_str.split('-')]
        start_date = parser.parse(new_period_value[0]).strftime('%Y-%m-%d')
        end_date = parser.parse(new_period_value[1]).strftime('%Y-%m-%d')
        dict_data['header'].update({'periode_mulai': start_date, 'periode_akhir': end_date})
        last_day_of_prev_month = datetime.strptime(start_date, '%Y-%m-%d').replace(day=1) - timedelta(days=1)
        start_day_of_prev_month = datetime.strptime(start_date, '%Y-%m-%d').replace(day=1) - timedelta(days=last_day_of_prev_month.day)
        last_day_of_prev_month = last_day_of_prev_month.strftime('%Y-%m-%d')
        start_day_of_prev_month = start_day_of_prev_month.strftime('%Y-%m-%d')
        
        # Search for Area Manager in the last row
        area_manager = sheet.cell(row=sheet.max_row, column=3).value or ''
        manager = None
        if area_manager:
            manager = self.env['hr.employee'].suspend_security().search([
                ('name', '=', str(area_manager)) ], limit=1)
            if manager:
                manager = manager.id

        dict_data['header'].update({
            'opex_avg': opex_value,
            'total_unit': total_unit_value,
            'total_cash': total_cash_value,
            'total_credit': total_credit_value,
            'total_net_margin': total_net_margin_value,
            'pbt_propose': pbt_propose_value
        })

        # cek apakah data transaksi sudah pernah dibuat sebelumnya
        # jika sudah, update data yg perlu di-update
        # jika belum, create transaksi baru
        # apakah ada kemungkinan pada 1 periode, terdapat lebih dari 1 header/transaksi?
        # jika ada case, periode masih satu rentang maka akan tetap dibuat transaksi baru. Misal A: periode 1 Jan 2021 - 31 Jan 2021, B: 10 Jan 2021 - 25 Jan 2021
        
        model_master_periode = self.env['tw.period.profit.before.tax']
        master_periode = model_master_periode.suspend_security().search([
            ('company_id', '=', company_id),
            ('start_date', '=', start_date),
            ('end_date', '=', end_date)
        ], limit=1)

        if not master_periode:
            raise Warning(f'Master periode {start_date} - {end_date} untuk branch {branch_obj.code} belum tersedia!')
        
        res_margin = self.env['tw.profit.before.tax']
        res_margin_line = self.env['tw.profit.before.tax.line']
        net_margin = res_margin.suspend_security().search([
            ('company_id', '=', company_id),
            ('start_date', '=', start_date),
            ('end_date', '=', end_date)
        ], limit=1, order='id DESC')
        net_margin_lm = res_margin.suspend_security().search([
            ('company_id', '=', company_id),
            ('start_date', '>=', start_day_of_prev_month),
            ('end_date', '<=', last_day_of_prev_month),
            ('state', '=', 'approved')
        ], limit=1, order='id DESC')

        fail = success = updated = 0
        fail_message = success_message = updated_message = ''
        fix_ttl_sales = 0

        fix_net_margin_salesman_cash = fix_net_margin_salesman_credit = fix_amount_net_margin_salesman = 0
        fix_net_margin_counter_cash = fix_net_margin_counter_credit = fix_amount_net_margin_counter = 0
        fix_net_margin_sco_cash = fix_net_margin_sco_credit = fix_amount_net_margin_sco = 0
        fix_all_net_margin_cash = fix_all_net_margin_credit = fix_all_net_margin_amount = 0

        # self.env.cr.execute(SQL("""
        #     SELECT series.name
        #     FROM product_category categ
        #     LEFT JOIN product_category segment ON segment.parent_id = categ.id
        #     LEFT JOIN product_category sub_segment ON sub_segment.parent_id = segment.id
        #     LEFT JOIN product_category series ON series.parent_id = sub_segment.id
        #     WHERE categ.name = 'Unit'
        # """))
        # product_series = self.env.cr.fetchall() 
        # series_list = [series[0] for series in product_series]
        auto_approved_header = True
        #loop untuk error checking data line
        line_ids = []
        vals = {}
        for row in range(15, sheet.max_row + 1):
            if row in (sheet.max_row - 1, sheet.max_row):
                continue
            elif not sheet.cell(row=row, column=2).value:
                continue
            else:
                series_name = sheet.cell(row=row, column=2).value
                year = sheet.cell(row=row, column=3).value
                amount_net_proposed_margin = sheet.cell(row=row, column=4).value
                lm_amount_net_proposed_margin = sheet.cell(row=row, column=5).value
                salesman_net_proposed_margin = sheet.cell(row=row, column=6).value
                lm_salesman_net_proposed_margin = sheet.cell(row=row, column=7).value
                counter_net_proposed_margin = sheet.cell(row=row, column=8).value
                lm_counter_net_proposed_margin = sheet.cell(row=row, column=9).value
                sco_net_proposed_margin = sheet.cell(row=row, column=10).value
                lm_sco_net_proposed_margin = sheet.cell(row=row, column=11).value
                description = sheet.cell(row=row, column=12).value
                lm_ttl_sales = sheet.cell(row=row, column=13).value
                lm_ttl_sales_cash = sheet.cell(row=row, column=14).value
                lm_ttl_sales_credit = sheet.cell(row=row, column=15).value
                ttl_sales = sheet.cell(row=row, column=16).value
                unit_cash_salesman = sheet.cell(row=row, column=17).value
                unit_cash_scounter = sheet.cell(row=row, column=18).value
                unit_cash_sco = sheet.cell(row=row, column=19).value
                amount_unit_cash = sheet.cell(row=row, column=20).value
                unit_credit_salesman = sheet.cell(row=row, column=21).value
                unit_credit_scounter = sheet.cell(row=row, column=22).value
                unit_credit_sco = sheet.cell(row=row, column=23).value
                amount_unit_credit = sheet.cell(row=row, column=24).value
                discount_cash_salesman = sheet.cell(row=row, column=25).value
                discount_cash_counter = sheet.cell(row=row, column=26).value
                discount_cash_sco = sheet.cell(row=row, column=27).value
                amount_discount_cash = sheet.cell(row=row, column=28).value
                discount_credit_salesman = sheet.cell(row=row, column=29).value
                discount_credit_counter = sheet.cell(row=row, column=30).value
                discount_credit_sco = sheet.cell(row=row, column=31).value
                amount_discount_credit = sheet.cell(row=row, column=32).value
                gp_bbn = sheet.cell(row=row, column=33).value
                gp_unit = sheet.cell(row=row, column=34).value
                ttl_gp = sheet.cell(row=row, column=35).value
                net_margin_salesman_cash = sheet.cell(row=row, column=36).value
                net_margin_salesman_credit = sheet.cell(row=row, column=37).value
                amount_net_margin_salesman = sheet.cell(row=row, column=38).value
                net_margin_counter_cash = sheet.cell(row=row, column=39).value
                net_margin_counter_credit = sheet.cell(row=row, column=40).value
                amount_net_margin_counter = sheet.cell(row=row, column=41).value
                net_margin_sco_cash = sheet.cell(row=row, column=42).value
                net_margin_sco_credit = sheet.cell(row=row, column=43).value
                amount_net_margin_sco = sheet.cell(row=row, column=44).value
                all_net_margin_cash = sheet.cell(row=row, column=45).value
                all_net_margin_credit = sheet.cell(row=row, column=46).value
                all_net_margin_amount = sheet.cell(row=row, column=47).value

                warning_note = []
                series = self.env['product.series'].search([('name', '=', str(series_name))], limit=1)
                if not series:
                    warning_note.append(f"Product series named {series_name} does not exist")

                if self.isinstance_check(amount_net_proposed_margin):
                    warning_note.append(f"Data Check ALL Position NET MARGIN/unit proposal is missing")
                
                if self.isinstance_check(lm_amount_net_proposed_margin):
                    warning_note.append(f"Data Check ALL Position NET MARGIN/unit Actual (LM) is missing")
                
                if self.isinstance_check(salesman_net_proposed_margin):
                    warning_note.append(f"Data Check Salesman NET MARGIN/unit proposal is missing")
                
                if self.isinstance_check(lm_salesman_net_proposed_margin):
                    warning_note.append(f"Data Check Salesman NET MARGIN/unit Actual (LM) is missing")
                
                if self.isinstance_check(counter_net_proposed_margin):
                    warning_note.append(f"Data Check Sales Counter NET MARGIN/unit proposal is missing")

                if self.isinstance_check(lm_counter_net_proposed_margin):
                    warning_note.append(f"Data Check Sales Counter NET MARGIN/unit Actual (LM) is missing")
                
                if self.isinstance_check(sco_net_proposed_margin):
                    warning_note.append(f"Data Check SCO NET MARGIN/unit proposal is missing")
                
                if self.isinstance_check(lm_sco_net_proposed_margin):
                    warning_note.append(f"Data Check SCO NET MARGIN/unit Actual (LM) is missing")
                
                if self.isinstance_check(description):
                    warning_note.append(f"Data Description is missing")

                if self.isinstance_check(lm_ttl_sales):
                    warning_note.append(f"Data TOTAL SALES LM is missing")
                
                if self.isinstance_check(lm_ttl_sales_cash):
                    warning_note.append(f"Data TOTAL SALES LM Cash is missing")
                
                if self.isinstance_check(lm_ttl_sales_credit):
                    warning_note.append(f"Data TOTAL SALES LM Credit is missing")
                
                if self.isinstance_check(ttl_sales):
                    warning_note.append(f"Data TOTAL SALES is missing")
                
                if self.isinstance_check(unit_cash_salesman):
                    warning_note.append(f"Data UNIT (CASH) SALESMAN is missing")
                
                if self.isinstance_check(unit_cash_scounter):
                    warning_note.append(f"Data UNIT (CASH) S.COUNTER is missing")

                if self.isinstance_check(unit_cash_sco):
                    warning_note.append(f"Data UNIT (CASH) SCO is missing")

                if self.isinstance_check(amount_unit_cash):
                    warning_note.append(f"Data TOTAL UNIT (CASH) is missing")
                
                if self.isinstance_check(unit_credit_salesman):
                    warning_note.append(f"Data UNIT (CREDIT) Salesman is missing")

                if self.isinstance_check(unit_credit_scounter):
                    warning_note.append(f"Data UNIT (CREDIT) S.COUNTER is missing")

                if self.isinstance_check(unit_credit_sco):
                    warning_note.append(f"Data UNIT (CREDIT) SCO is missing")
                
                if self.isinstance_check(amount_unit_credit):
                    warning_note.append(f"Data TOTAL UNIT (CREDIT) is missing")
                
                if self.isinstance_check(discount_cash_salesman):
                    warning_note.append(f"Data DISCOUNT (CASH) (dealer+HC burden) Salesman is missing")
                
                if self.isinstance_check(discount_cash_counter):
                    warning_note.append(f"Data DISCOUNT (CASH) (dealer+HC burden) S.COUNTER is missing")
                
                if self.isinstance_check(discount_cash_sco):
                    warning_note.append(f"Data DISCOUNT (CASH) (dealer+HC burden) SCO is missing")
                
                if self.isinstance_check(amount_discount_cash):
                    warning_note.append(f"Data TOTAL DISCOUNT (CASH) is missing")
                
                if self.isinstance_check(discount_credit_salesman):
                    warning_note.append(f"Data DISCOUNT (CREDIT) (dealer+HC burden) Salesman is missing")
                
                if self.isinstance_check(discount_credit_counter):
                    warning_note.append(f"Data DISCOUNT (CREDIT) (dealer+HC burden) S.COUNTER is missing")
                
                if self.isinstance_check(discount_credit_sco):
                    warning_note.append(f"Data DISCOUNT (CREDIT) (dealer+HC burden) SCO is missing")
                
                if self.isinstance_check(amount_discount_credit):
                    warning_note.append(f"Data TOTAL DISCOUNT (CREDIT) is missing")

                if self.isinstance_check(gp_bbn):
                    warning_note.append(f"Data GROSS PROFIT BBN is missing")

                if self.isinstance_check(gp_unit):
                    warning_note.append(f"Data GROSS PROFIT UNIT is missing")

                if self.isinstance_check(ttl_gp):
                    warning_note.append(f"Data TOTAL GP is missing")

                if self.isinstance_check(net_margin_salesman_cash):
                    warning_note.append(f"Data NET MARGIN Salesman CASH is missing")
                
                if self.isinstance_check(net_margin_salesman_credit):
                    warning_note.append(f"Data NET MARGIN Salesman CREDIT is missing")
                
                if self.isinstance_check(amount_net_margin_salesman):
                    warning_note.append(f"Data NET MARGIN Salesman TOTAL is missing")
                
                if self.isinstance_check(net_margin_counter_cash):
                    warning_note.append(f"Data NET MARGIN SC CASH is missing")
                
                if self.isinstance_check(net_margin_counter_credit):
                    warning_note.append(f"Data NET MARGIN SC CREDIT is missing")
                
                if self.isinstance_check(amount_net_margin_counter):
                    warning_note.append(f"Data NET MARGIN SC TOTAL is missing")
                
                if self.isinstance_check(net_margin_sco_cash):
                    warning_note.append(f"Data NET MARGIN SCO CASH is missing")
                
                if self.isinstance_check(net_margin_sco_credit):
                    warning_note.append(f"Data NET MARGIN SCO CREDIT is missing")
                
                if self.isinstance_check(amount_net_margin_sco):
                    warning_note.append(f"Data NET MARGIN SCO TOTAL is missing")
                
                if self.isinstance_check(all_net_margin_cash):
                    warning_note.append(f"Data NET MARGIN ALL CASH is missing")
                
                if self.isinstance_check(all_net_margin_credit):
                    warning_note.append(f"Data NET MARGIN ALL CREDIT is missing")
                
                if self.isinstance_check(all_net_margin_amount):
                    warning_note.append(f"Data NET MARGIN ALL TOTAL is missing")

                if warning_note:
                    fail += 1
                    warning = ', '.join(warning_note)
                    fail_message += f"Line number {row + 1} {warning}.\n"
                    continue
                else:
                    fix_ttl_sales += ttl_sales
                    
                    fix_net_margin_salesman_cash += net_margin_salesman_cash
                    fix_net_margin_salesman_credit += net_margin_salesman_credit
                    fix_amount_net_margin_salesman += amount_net_margin_salesman                    
                    
                    fix_net_margin_counter_cash += net_margin_counter_cash
                    fix_net_margin_counter_credit += net_margin_counter_credit
                    fix_amount_net_margin_counter += amount_net_margin_counter
                    
                    fix_net_margin_sco_cash += net_margin_sco_cash
                    fix_net_margin_sco_credit += net_margin_sco_credit
                    fix_amount_net_margin_sco += amount_net_margin_sco
                    
                    fix_all_net_margin_cash += all_net_margin_cash
                    fix_all_net_margin_credit += all_net_margin_credit
                    fix_all_net_margin_amount += all_net_margin_amount
                                
                    # jika header transaksi sudah ada
                    line_vals = {
                        'series_id': series.id,
                        'year': year,
                        'amount_net_proposed_margin': int(amount_net_proposed_margin),
                        'lm_amount_net_proposed_margin': int(lm_amount_net_proposed_margin),
                        'salesman_net_proposed_margin': int(salesman_net_proposed_margin),
                        'lm_salesman_net_proposed_margin': int(lm_salesman_net_proposed_margin),
                        'counter_net_proposed_margin': int(counter_net_proposed_margin),
                        'lm_counter_net_proposed_margin': int(lm_counter_net_proposed_margin),
                        'sco_net_proposed_margin': int(sco_net_proposed_margin),
                        'lm_sco_net_proposed_margin': int(lm_sco_net_proposed_margin),
                        'description': 'cek' if description == 'CEK' else 'ok',
                        'lm_ttl_sales': int(lm_ttl_sales),
                        'lm_ttl_sales_cash': int(lm_ttl_sales_cash),
                        'lm_ttl_sales_credit': int(lm_ttl_sales_credit),
                        'ttl_sales': int(ttl_sales),
                        'unit_cash_salesman': int(unit_cash_salesman),
                        'unit_cash_scounter': int(unit_cash_scounter),
                        'unit_cash_sco': int(unit_cash_sco),
                        'amount_unit_cash': int(amount_unit_cash),
                        'unit_credit_salesman': int(unit_credit_salesman),
                        'unit_credit_scounter': int(unit_credit_scounter),
                        'unit_credit_sco': int(unit_credit_sco),
                        'amount_unit_credit': int(amount_unit_credit),
                        'discount_cash_salesman': int(discount_cash_salesman),
                        'discount_cash_counter': int(discount_cash_counter),
                        'discount_cash_sco': int(discount_cash_sco),
                        'amount_discount_cash': int(amount_discount_cash),
                        'discount_credit_salesman': int(discount_credit_salesman),
                        'discount_credit_counter': int(discount_credit_counter),
                        'discount_credit_sco': int(discount_credit_sco),
                        'amount_discount_credit': int(amount_discount_credit),
                        'gp_bbn': int(gp_bbn),
                        'gp_unit': int(gp_unit),
                        'ttl_gp': int(ttl_gp),
                        'net_margin_salesman_cash': int(net_margin_salesman_cash),
                        'net_margin_salesman_credit': int(net_margin_salesman_credit),
                        'amount_net_margin_salesman': int(amount_net_margin_salesman),
                        'net_margin_counter_cash': int(net_margin_counter_cash),
                        'net_margin_counter_credit': int(net_margin_counter_credit),
                        'amount_net_margin_counter': int(amount_net_margin_counter),
                        'net_margin_sco_cash': int(net_margin_sco_cash),
                        'net_margin_sco_credit': int(net_margin_sco_credit),
                        'amount_net_margin_sco': int(amount_net_margin_sco),
                        'all_net_margin_cash': int(all_net_margin_cash),
                        'all_net_margin_credit': int(all_net_margin_credit),
                        'all_net_margin_amount': int(all_net_margin_amount),
                        'state': 'draft' if description == 'CEK' else 'approved',
                        'approve_uid': self._uid if description == 'CEK' else False,
                        'approve_date': self._get_default_datetime() if description == 'CEK' else False
                    }
                    if line_vals['state'] == 'draft' or description == 'CEK':
                        auto_approved_header = False

                    if net_margin and net_margin.state == 'draft':
                        margin_line = res_margin_line.suspend_security().search([
                            ('net_margin_id', '=', net_margin.id),
                            ('series_motor', '=', series.id),
                            ('year', '=', year)], limit=1)
                        
                        if margin_line.state == 'approved':
                            continue
                        
                        if margin_line:
                            line_vals['ttl_sales'] = line_vals.get('ttl_sales', 0) + margin_line.ttl_sales
                            line_ids.append(Command.update(margin_line.id, line_vals))
                            updated += 1
                            updated_message += f'Data baris ke {row+1} berhasil di update!\n'
                        else:
                            line_ids.append(Command.create(line_vals))
                            success += 1
                            success_message += f'Data baris ke {row+1} berhasil di tambahkan!\n'
                    # jika header transaksi belum ada, create data transaksi baru
                    else:
                        line_vals.update({
                            'company_id': dict_data['header']['kode_cabang'],
                            'start_date': dict_data['header']['periode_mulai'],
                            'end_date': dict_data['header']['periode_akhir']
                        })
                        line_ids.append(Command.create(line_vals))
                        success += 1
                        success_message += f'Data transaksi baru dan data baris ke {row+1} berhasil di tambahkan!\n'
        
        if fail_message:
            raise Warning(_(fail_message))
        
        vals.update({
            'company_id': master_periode.company_id.id,
            'start_date': master_periode.start_date,
            'end_date': master_periode.end_date,
            'opex_avg': master_periode.opex_avg,
            'total_unit_lm': master_periode.total_unit_lm or 0,
            'total_net_margin_lm': master_periode.total_net_margin_lm or 0,
            'total_refund_lm': master_periode.refund_lm or 0,
            'pbt_propose_lm': master_periode.pbt_propose_lm or 0,
            'area_manager_id': manager,
            'profit_before_tax_line_ids': line_ids
        })
        
        if net_margin and net_margin.state == 'draft':
            net_margin.suspend_security().write(vals)
        else:
            net_margin = res_margin.suspend_security().create(vals)
        
        # auto generate master target margin if all state lines are approved
        if auto_approved_header:
            net_margin.action_request_approval()
            net_margin.action_approval()
            net_margin.action_confirm()
                    
        messages = ""
        if fail:
            messages += f"{fail} Data TIDAK berhasil di upload! Alasan Gagal : \n{fail_message}"
            raise Warning(messages)
        
        if success:
            messages += f'{success} Data Baru Berhasil di Buat'
        if updated:
            if success:
                messages += " Dan"
            messages += f'{updated} Data Berhasil di Update'

        self.messages = messages
        self.state_x = 'get'
        
        list_id = self.env.ref('tw_profit_before_tax.view_tw_profit_before_tax_list').id
        form_id = self.env.ref('tw_profit_before_tax.view_tw_profit_before_tax_form').id
        
        return {
            'type': 'ir.actions.act_window',
            'name': 'Input Net Margin',
            'view_mode': 'list,form',
            'res_model': 'tw.profit.before.tax',
            'views': [(list_id, 'list'), (form_id, 'form')],
            'context': { 'readonly_by_pass': 1 }
        }