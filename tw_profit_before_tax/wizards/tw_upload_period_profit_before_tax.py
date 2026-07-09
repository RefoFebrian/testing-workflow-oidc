
import base64
import os
import calendar
import logging
import pandas as pd
import openpyxl

from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from io import BytesIO


from odoo import models, fields, api, _
from odoo.exceptions import UserError as Warning
from odoo.tools import SQL
from odoo.tools.misc import formatLang, file_path as file_path_util

_logger = logging.getLogger(__name__)
_logger.setLevel(logging.INFO)


class UploadPeriodProfitBeforeTax(models.TransientModel):
    _name = "tw.upload.period.profit.before.tax"
    _description = "Upload Period Profit Before Tax"

    def _get_default_date(self):
        return self.env['res.company'].get_default_date()

    @api.model
    def _get_default_datetime(self):
        return self.env['res.company']._get_default_datetime_plus_7()

    @api.model
    def _get_first_date_of_the_month(self):
        return self.env['res.company'].get_default_date().replace(day=1)

    @api.model
    def _get_last_date_of_the_month(self):
        now = self.env['res.company'].get_default_date()
        month = now.month
        year = now.year
        last_date = calendar.monthrange(year, month)[1]
        return datetime(year, month, last_date)

    file = fields.Binary('File')
    date = fields.Date('Tanggal', readonly=True, default=_get_default_date)
    data_x = fields.Binary('File Output', readonly=True)
    state_x = fields.Selection([
        ('choose', 'choose'), ('get', 'get')],
        default=lambda self: self._context.get('default_state_x', 'choose')
    )
    messages = fields.Text("message", readonly=True)

    def action_download_format_file(self):
        today = date.today()
        lomonth = today.replace(day=1) - relativedelta(days=1)
        lmonth = lomonth.replace(day=1)
        query = """
            SELECT wb.code as code_branch
                , SUM(dsol.product_uom_qty) AS qty
                , SUM(dsol.net_margin) AS gp_total
                , SUM(dsol.finco_incentive) AS insentif_finco
            from tw_dealer_sale_order dso 
            LEFT JOIN tw_dealer_sale_order_line dsol on dsol.order_id = dso.id
            LEFT JOIN account_tax_tw_dealer_sale_order_line_rel dsot ON dsot.tw_dealer_sale_order_line_id = dsol.id
            LEFT JOIN account_tax at ON at.id = dsot.account_tax_id
            LEFT JOIN product_product product ON dsol.product_id = product.id
            LEFT JOIN ( 
                SELECT order_line_id,
                    SUM(COALESCE(amount_finco, 0)) AS ps_finco,
                    SUM(COALESCE(amount_ahm, 0)) AS ps_ahm,
                    SUM(COALESCE(amount_md, 0)) AS ps_md,
                    SUM(COALESCE(amount_dealer, 0)) AS ps_dealer,
                    SUM(COALESCE(amount_others, 0)) AS ps_others, 
                    SUM(COALESCE(discount_amount, 0)) AS discount,
                    SUM(COALESCE(discount_customer, 0)) AS discount_pelanggan,
                    SUM(COALESCE(amount_diff_md, (amount_ahm::float + amount_md) - discount_customer)) AS ps_md_diff,
                    SUM(COALESCE(amount_diff_finco, amount_finco::float - discount_customer)) AS ps_finco_diff
                FROM tw_dealer_sale_order_line_program 
                GROUP BY order_line_id 
            ) dsol_disc ON dsol_disc.order_line_id = dsol.id
            LEFT JOIN (
                SELECT order_line_id,
                    sum(direct_gift_dealer) AS total_brg_bonus
                FROM tw_dealer_sale_order_line_direct_gift
                GROUP BY order_line_id
            ) dsol_brg ON dsol_brg.order_line_id = dsol.id
            LEFT JOIN res_company wb on wb.id = dso.company_id
            WHERE dso.state IN ('sale', 'done')
            AND dso.date_order BETWEEN %(start_date)s AND %(end_date)s
            AND dsol.item_type = 'main'
            GROUP BY wb.id
        """

        params = {'start_date': lmonth, 'end_date': lomonth}
        try:
            self.env.cr.execute(SQL(query, **params))
            results = self.env.cr.dictfetchall()
        except Exception as e:
            raise Warning(_(str(e)))

        if not results:
            raise Warning(_("No sales data found!"))

        # use this on production or tes server!
        format = self.env['tw.format.upload'].suspend_security().search([
            ('name', '=', 'master period pbt'),
            ('active', '=', True)], limit=1)
        if format:
            if not format.filename_upload_format:
                raise Warning(_("Filename for 'master period pbt' format is not configured. Please contact the Helpdesk."))

            path = self.env['tw.config.files'].search([('name', '=', 'FORMAT-UPLOAD')])
            if not path:
                raise Warning(_("Configuration for 'FORMAT-UPLOAD' not found. Please ensure the configuration exists in the system."))
            
            full_path = f'{path.local_path}/{format.filename_upload_format}'
        
        elif not format:
            # Fallback: serve from module data/ folder
            filename = 'format_master_periode_sisa_margin.xlsx'
            full_path = file_path_util(f'tw_profit_before_tax/data/example/{filename}')
            if not full_path or not os.path.exists(full_path):
                raise UserError(_("Format template belum tersedia. Silakan hubungi tim IT."))

        excel_data = pd.read_excel(
            full_path, header=None, usecols=[0, 1, 2, 3, 4, 5])

        today = datetime.today().strftime('%d-%m-%Y')
        filename = f'FORMAT MASTER PERIODE INPUT SISA MARGIN {today}.xlsx'

        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        excel_data.to_excel(writer, sheet_name='Sheet1',
                            index=False, header=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        format1 = workbook.add_format({
            'bg_color': '#FFFF00',
            'bold': 1,
            'align': 'center',
            'font_color': '#000000',
            'border': 1
        })
        format2 = workbook.add_format({
            'bg_color': '#FFFFFF',
            'align': 'left',
            'font_color': '#000000',
            'border': 1
        })
        format3 = workbook.add_format({'bold': 1})
        format4 = workbook.add_format(
            {'align': 'center', 'font_color': '#000000'})
        format5 = workbook.add_format({
            'align': 'right',
            'num_format': 'Rp#,##0.00',
            'border': 1
        })
        format6 = workbook.add_format({'align': 'right', 'border': 1})

        worksheet.set_column('B1:B1', 20)
        worksheet.set_column('C1:C1', 25)
        worksheet.set_column('D1:D1', 20)
        worksheet.set_column('E1:E1', 25)
        worksheet.set_column('F1:F1', 25)
        worksheet.set_column('G1:G1', 25)

        worksheet.write('B1', '=TODAY()', format4)
        worksheet.write('B2', 'PERIODE', format3)
        worksheet.write('C1', '=EOMONTH(B1, -1)+1', format4)
        worksheet.write('C2', '=Sheet1!E1&" - "&Sheet1!F1', format4)
        worksheet.write('D1', '=EOMONTH(C1, 0)', format4)
        worksheet.write('E1', '=TEXT(C1, "dd mmm yyyy")', format4)
        worksheet.write('F1', '=TEXT(D1, "dd mmm yyyy")', format4)

        worksheet.write('B4', excel_data.iat[3, 1], format1)
        worksheet.write('C4', excel_data.iat[3, 2], format1)
        worksheet.write('D4', excel_data.iat[3, 3], format1)
        worksheet.write('E4', excel_data.iat[3, 4], format1)
        worksheet.write('F4', excel_data.iat[3, 5], format1)
        worksheet.write('G4', 'REFUND LM', format1)

        row = 4
        for record in results:
            worksheet.write(f'B{row+1}', record.get('code_branch'), format2)
            worksheet.write(f'C{row+1}', 0, format5)
            worksheet.write(f'D{row+1}', record.get('qty'), format6)
            worksheet.write(f'E{row+1}', record.get('gp_total'), format5)
            worksheet.write(f'F{row+1}', 0, format5)
            worksheet.write(f'G{row+1}', record.get('insentif_finco'), format5)
            row += 1

        writer.close()
        output.seek(0)
        out = base64.b64encode(output.getvalue())
        self.data_x = out

        return {
            'type': 'ir.actions.act_url',
            'name': 'contract',
            'url': f'/web/content/tw.upload.period.profit.before.tax/{self.id}/data_x/{filename}?download=true'
        }

    def isinstance_check(self, value):
        if isinstance(value, str) and value == '':
            return True
        else:
            return False

    def action_import(self):
        if not self.file:
            raise Warning('Silakan input file terlebih dahulu.')

        data = base64.b64decode(self.file)
        excel = openpyxl.load_workbook(BytesIO(data), data_only=True)
        sheet = excel.worksheets[0]

        warning_header = ''

        # Search for the period value dynamically
        period_value = sheet.cell(row=2, column=3).value
        # If C2 is empty, look for the 'PERIODE' label and take the value next to it
        if not period_value or self.isinstance_check(period_value):
            for r in range(1, 5):
                for c in range(1, 5):
                    cell_val = sheet.cell(row=r, column=c).value
                    if cell_val and "PERIODE" in str(cell_val).upper():
                        period_value = sheet.cell(row=r, column=c+1).value
                        break
                if period_value:
                    break

        if not period_value or self.isinstance_check(period_value):
            warning_header += 'Data PERIODE tidak ditemukan (seharusnya di sebelah label PERIODE atau di sel C2)\n'

        if warning_header:
            raise Warning(
                'Tidak dapat diproses, karena ada kesalahan.\nDetail:\n{}'.format(warning_header))

        # Convert to string and handle formatting (prevents AttributeError: 'int' object has no attribute 'split')
        period_str = str(period_value).strip() if period_value else ""
        if '-' not in period_str:
            raise Warning(_("Format periode '%s' salah. Harap gunakan format 'Tgl Awal - Tgl Akhir' (contoh: 01 Jan 2026 - 31 Jan 2026).") % period_str)

        new_period_value = [item.strip() for item in period_str.split('-')]
        start_period = new_period_value[0]
        end_period = new_period_value[1]
        start_date = datetime.strptime(start_period, "%d %b %Y").strftime('%Y-%m-%d')
        end_date = datetime.strptime(end_period, "%d %b %Y").strftime('%Y-%m-%d')

        fail_message = fail_kode_cabang = ''
        fail = 0

        for row in range(5, sheet.max_row + 1):
            vals = {}
            warning_note = ''
            warning_kode_cabang = ''
            branch_obj = self.env['res.company'].search([('code', '=', sheet.cell(row=row, column=2).value)], limit=1)
            opex_avg = sheet.cell(row=row, column=3).value
            total_unit_lm = sheet.cell(row=row, column=4).value
            total_net_margin_lm = sheet.cell(row=row, column=5).value
            pbt_propose_lm = sheet.cell(row=row, column=6).value
            refund_lm = sheet.cell(row=row, column=7).value

            if not branch_obj:
                warning_kode_cabang += str(sheet.cell(row=row, column=2).value) + ', '
            if self.isinstance_check(opex_avg):
                warning_note += f'Baris ke {row+1} data OPEX AVG tidak ada\n'
            if self.isinstance_check(total_unit_lm):
                warning_note += f'Baris ke {row+1} data Total Unit LM tidak ada\n'
            if self.isinstance_check(total_net_margin_lm):
                warning_note += f'Baris ke {row+1} data Total Net Margin LM tidak ada\n'
            if self.isinstance_check(pbt_propose_lm):
                warning_note += f'Baris ke {row+1} data PBT Propose LM tidak ada\n'
            if self.isinstance_check(refund_lm):
                warning_note += f'Baris ke {row+1} data Total Refund LM tidak ada\n'

            if warning_note or warning_kode_cabang:
                if warning_note:
                    fail_message += warning_note + '\n'
                if warning_kode_cabang:
                    fail_kode_cabang += warning_kode_cabang
                fail += 1
                continue
            else:
                master_object = self.env['tw.period.profit.before.tax']
                period_margin = master_object.suspend_security().search([
                    ('company_id', '=', branch_obj.id),
                    ('start_date', '=', start_date),
                    ('end_date', '=', end_date)
                ], limit=1)

                if period_margin:
                    if period_margin.opex_avg != int(opex_avg):
                        vals.update({'opex_avg': opex_avg})
                    if period_margin.total_unit_lm != int(total_unit_lm):
                        vals.update({'total_unit_lm': total_unit_lm})
                    if period_margin.total_net_margin_lm != int(total_net_margin_lm):
                        vals.update(
                            {'total_net_margin_lm': total_net_margin_lm})
                    if period_margin.pbt_propose_lm != int(pbt_propose_lm):
                        vals.update({'pbt_propose_lm': pbt_propose_lm})
                    if period_margin.refund_lm != int(refund_lm):
                        vals.update({'refund_lm': refund_lm})
                    if vals:
                        period_margin.write(vals)
                else:
                    vals.update({
                        'company_id': branch_obj.id,
                        'start_date': start_date,
                        'end_date': end_date,
                        'opex_avg': opex_avg,
                        'total_unit_lm': total_unit_lm,
                        'total_net_margin_lm': total_net_margin_lm,
                        'pbt_propose_lm': pbt_propose_lm,
                        'refund_lm': refund_lm
                    })
                    master_object.create(vals)

        messages = ""
        if fail_kode_cabang:
            messages += "\n Kode Cabang Berikut:\n{}\nTidak ditemukan di master.".format(
                fail_kode_cabang[:-2])
        if fail_message:
            messages += "\n {} Data TIDAK berhasil di upload!\
                \n\n Alasan Gagal : \n {}".format(fail, fail_message)

        if not messages:
            messages += 'Import Sukses, Tidak Ada Kesalahan Untuk Seluruh Data Yang Di Import.'
        self.messages = messages
        self.state_x = 'get'

        submenu_name = 'Upload Period Profit Before Tax'
        res_model = 'tw.upload.period.profit.before.tax'
        form_id = self.env.ref(
            'tw_profit_before_tax.tw_upload_period_profit_before_tax_wizard').id

        result = {
            'type': 'ir.actions.act_window',
            'name': (submenu_name),
            'res_model': res_model,
            'views': [(form_id, 'form')],
            'view_id': form_id,
            'target': 'new',
            'res_id': self.ids[0]
        }

        return result
