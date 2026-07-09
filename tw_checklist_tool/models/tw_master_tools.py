# 1: imports of python lib
import base64
import datetime
from io import BytesIO
import openpyxl
import xlrd
from dateutil.relativedelta import relativedelta
import logging

_logger = logging.getLogger(__name__)

# 2: import of known third party lib

# 3:  imports of odoo
from odoo import models, api, fields
from odoo.exceptions import ValidationError, UserError

# 4:  imports from odoo modules

# 5: local imports

# 6: Import of unknown third party lib

COLS = {
    'Product Code': 3,
    'Qty': 5,
}

class TwMasterTools(models.Model):
    _name = "tw.master.tools"
    _description = "Master Tools"

    # 7: defaults methods
    @api.model
    def _get_default_branch(self):
        company_ids = False
        company_ids = self.env.user.company_ids
        if company_ids and len(company_ids) == 1:
            return company_ids[0].id
        return False

    # 8: fields
    name = fields.Char('Name', required=True)
    file = fields.Binary('File')
    filename = fields.Char('Filename')

    # 9: relation fields
    company_id = fields.Many2one('res.company', string="Branch")
    category_master_tool_id = fields.Many2one('tw.selection', "Category", domain=[('type', '=', 'MasterToolCategory')])
    pic_id = fields.Many2one('hr.employee', string="PIC")
    location_id = fields.Many2one('tw.selection', string="Location", domain="[('type', '=', 'MasterToolsLocationType')]")

    tw_master_tools_line_ids = fields.One2many('tw.master.tools.line', 'master_tools_id', string="Master Tools Line")

    # 10: constraints & sql constraints
    @api.constrains('company_id', 'pic_id', 'category_master_tool_id', 'location_id')
    def _check_existing_master(self):
        for record in self:
            if record.company_id and record.pic_id and record.category_master_tool_id and record.location_id:
                existing = self.search([
                    ('id', '!=', record.id),
                    ('company_id', '=', record.company_id.id),
                    ('pic_id', '=', record.pic_id.id),
                    ('category_master_tool_id', '=', record.category_master_tool_id.id),
                    ('location_id', '=', record.location_id.id),
                ], limit=1)

                if existing:
                    raise ValidationError(
                        "Master Tools dengan Branch {branch}, PIC {pic}, Location {location}, dan Category {categ} sudah ada!".format(
                            branch=record.company_id.name,
                            pic=record.pic_id.name,
                            location=record.location_id.name,
                            categ=record.category_master_tool_id.name
                        )
                    )

    # 11: compute/depends & on change methods
    @api.onchange('company_id')
    def onchange_employee(self):
        self.pic_id = False
        ids = []
        if self.company_id:
            jobs = self.env['hr.job'].search([('name', 'in', ('MECHANIC', 'MEKANIK MITRA', 'PARTMAN', 'FRONT DESK', 'FRONT DESK (PARTMAN)', 'MECHANIC HEAD', 'SERVICE ADVISOR', 'WORKSHOP HEAD'))])
            empl = self.env['hr.employee'].sudo().search([
                ('company_id', '=', self.company_id.id),
                ('job_id', 'in', [j.id for j in jobs]),
                ('working_end_date', '=', False), ('active', '=', True)])
            ids = [e.id for e in empl]
        domain = {'pic_id': [('id', 'in', ids)]}
        return {'domain': domain}

    # 12: override methods
    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('file'):
                files = vals.get('file')

                filename_upload_tokens = str(vals.get('filename')).split('.')
                now = (datetime.datetime.today() + relativedelta(hours=7)).strftime('-%Y-%m-%d_%H_%M_%S_%f')
                filename = str('tw_checklist_tool-') + now + '.' + filename_upload_tokens[len(filename_upload_tokens) - 1]

                self.env['tw.config.files'].suspend_security().upload_file(filename, files)
                vals['file'] = False
                vals['filename'] = filename

            vals['name'] = "Master Tool"

        return super(TwMasterTools, self).create(vals_list)

    def write(self, vals):
        if vals.get('file'):
            files = vals.get('file')
            filename_upload_tokens = str(vals.get('filename')).split('.')
            now = (datetime.datetime.today() + relativedelta(hours=7)).strftime('-%Y-%m-%d_%H_%M_%S_%f')
            filename = str('tw_checklist_tool-') + now + '.' + filename_upload_tokens[len(filename_upload_tokens) - 1]

            self.env['tw.config.files'].suspend_security().upload_file(filename, files)
            vals['file'] = False
            vals['filename'] = filename

        return super(TwMasterTools, self).write(vals)

    def unlink(self):
        if self:
            raise UserError('Perhatian! Master Tools tidak bisa dihapus!')
        return super(TwMasterTools, self).unlink()

    # 13: action methods
    def action_import_tools_line(self):
        if not self.filename:
            raise UserError("Perhatian!\nSilakan Upload File Excel terlebih dahulu!")

        data_file = self.env['tw.config.files'].suspend_security().get_file(self.filename)
        data = base64.b64decode(data_file)
        lines = []
        fail = update = success = 0
        fail_message = update_msg = success_message = ''
        existing = {line.product_id.id: line for line in self.tw_master_tools_line_ids}

        ext = self.filename.split('.')[-1].lower()
        if ext == 'xls':
            excel = xlrd.open_workbook(file_contents=data)
            sheet = excel.sheet_by_index(0)
        elif ext == 'xlsx':
            excel = openpyxl.load_workbook(BytesIO(data), data_only=True)
            sheet = excel.active
        else:
            raise UserError('TIDAK dapat upload file dengan ekstensi %s' % ext)

        iteration = range(1, sheet.nrows) if ext == 'xls' else range(1, sheet.max_row)
        x = 1 if ext == 'xlsx' else 0
        done_product = []
        for i in iteration:
            row = {
                'Product Code': self._read_cell_value(sheet.cell(i + x, COLS['Product Code'] + x).value),
                'Qty': self._read_cell_value(sheet.cell(i + x, COLS['Qty'] + x).value, expected_type='int'),
            }

            # If all row is None skip the iteration
            if any(v is None or str(v).strip() == '' for v in row.values()):
                fail += 1
                fail_message += 'Baris ke-{row}, Memiliki Data yang Kosong. \n\n'.format(row=(i))
                done_product.append(row['Product Code'])
                continue

            product_obj = self.env['product.product'].search(['|', ('default_code', '=', row['Product Code']), ('name', '=', row['Product Code'])], limit=1)
            if not product_obj:
                fail += 1
                fail_message += 'Baris ke-{row}, Product Code [{product_code}] TIDAK ditemukan. \n\n'.format(row=(i), product_code=row['Product Code'])
                done_product.append(row['Product Code'])
                continue

            if row['Product Code'] in done_product:
                fail += 1
                fail_message += 'Baris ke-{row}, Duplikasi Product Code [{product_code}] pada Format Upload. \n\n'.format(
                    row=(i), product_code=row['Product Code'])
                continue

            qty = row['Qty']
            if qty < 1:
                fail += 1
                fail_message += 'Baris ke-{row}, Qty tidak boleh bernilai 0 atau menggunakan huruf. \n\n'.format(
                    row=(i))
                done_product.append(row['Product Code'])
                continue

            data_line = existing.get(product_obj.id)
            if data_line:
                # Update existing line
                if data_line.qty_tool != qty:
                    update += 1
                    update_msg += "Baris ke-%s berhasil di‐update: Qty Tools %s dari %s → %s.\n" % (i, str(product_obj.name_template), data_line.qty_tool, qty)
                    data_line.qty_tool = qty
                    done_product.append(row['Product Code'])
            else:
                # new line entry
                success += 1
                success_message += "Baris ke-%d berhasil diimport.\n" % (i)
                done_product.append(row['Product Code'])
                lines.append((0, 0, {
                    'product_id': product_obj.id,
                    'qty_tool': row['Qty'],
                }))

        if lines:
            self.tw_master_tools_line_ids = lines

        messages = ""

        if fail > 0:
            messages += "{fail} Data TIDAK berhasil di upload! Alasan Gagal : \n{fail_message}".format(
                fail=fail,
                fail_message=fail_message
            )

        if success > 0:
            messages += '\n{success} Data Baru Berhasil di Buat'.format(success=success)

        if update > 0:
            messages += '\n{update} Data Berhasil di Update'.format(update=update)

        self.filename = False
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'tw.upload.message.wizard',
            'view_mode': 'form',
            'view_id': self.env.ref('tw_checklist_tool.view_tw_import_result_message_wizard').id,
            'target': 'new',
            'context': {'message': messages},
        }

    def _read_cell_value(self, value, force_zero_padding=False, use_fraction=False, expected_type='str'):
        if value is None:
            return ''

        if use_fraction:
            value = str(value).replace('.0','').replace(',0','')
            try:
                return float(value)
            except ValueError:
                return value  # fallback

        if isinstance(value, float) and force_zero_padding:
            # Phone number, codes that must keep leading zero
            value = '0' + str(int(value))

        try:
            value = str(value) if not isinstance(value, str) else value
        except UnicodeEncodeError as err:
            _logger.error(err.args)
            value = str(value.encode('utf-8'))

        # Trim leading/trailing spaces
        value = value.strip()

        if expected_type == 'int':
            try:
                return int(float(value))  # handles '2.0', '3', or even '3.00'
            except (ValueError, TypeError):
                return 0

        return value

    def action_download_format_file(self):
        format = self.env['tw.format.upload'].sudo().search(
            [('name', '=', 'upload master tools'), ('active', '=', True)],
            limit=1
        )

        if not format or not format.file_format_show:
            raise UserError('Maaf, format belum tersedia. Silakan hubungi Helpdesk.')

        return {
            'type': 'ir.actions.act_url',
            'url': (
                       '/web/content'
                       '?model=tw.format.upload'
                       '&id=%s'
                       '&field=file_format_show'
                       '&filename_field=filename_upload_format'
                       '&download=true'
                   ) % format.id,
            'target': 'self',
        }

    # 14: private methods